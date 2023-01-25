#!/bin/python3

""" A script wich transfers data from an excel-file to an *.ics and
send the result back via E-mail """



# native
import os
import re
import time
import datetime as dt
import smtplib
import ssl
import email
from email import header
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# installed
import Levenshtein
import openpyxl

# own
from modules import icsgenerator as ics

# constant (filepaths)

#Paths PC
OWN_PATH: str = "/home/XXX/XXX"
MAILBOX_PATH: str = "/home/XXX/XXX/data/"
FAULTY_MAILS: str = "/home/XXX/XXX/data/"

SMTP, EMAILADRESS, SERVICEADRESS, PASSWORD = "smtp.XXX.xx", "email@XXX.xx", "service@XXX.xx", "XXX"

def check_mailbox()-> list:
    """checks for new mails in the MAILBOX_PATH

    Returns: list
        mailfiles -- filenames of the unseen mails"""

    mailbox : list = os.listdir(MAILBOX_PATH)   # list of all mailfiles in folder

    for file in mailbox:
        if "S" in file.split(",")[-1:][0]:      # the "S" in the last part stands for seen
            mailbox.remove(file)

    return mailbox

def read_mail(mail:str)-> tuple:
    """checks if it's the right subject and extract the content of interest.

    Arguments:
        mail -- filename of the mail

    Returns: tuple
        sender -- email-adress from sender
        employee_name -- searched name
        filename -- filename of the attachment
    """

    def check_subject(subject: str)-> bool:
        """checks if a part of the subject is in the list for accepted subjects

        Arguments:
            subject -- the subject

        Returns:
            bool -- True if subject suits
        """

        accepted_subjects : list = ["schichtplan", "dienstplan"]

        if subject.isascii():
            subject : list = subject.strip(" ").split()

            for string in subject:
                for acc_subject in accepted_subjects:
                    if string.lower() in acc_subject:
                        return True
        return False

    def parse_plain(): # ToDo
        pass

    def parse_html_body(line: str):
        """parse the html line into a readable string and exclude "" and "\&nbsp;"

        Arguments:
            line -- line of the body

        Returns:
            False or the first line with logical content of the body
        """
        edited_line = re.sub(r"\<[^<>]*\>", "*", line).split("*")   # replace html-code from line
        for string in edited_line:
            if string not in ("", "&nbsp;"):
                employee_name: str = string
                return employee_name

        return False

    def get_content(message: email)-> tuple:
        """extracts the employeename and filename (attachment) from message

            walks over the multipart message until finds the first written line in body
            and until finds the filename from *.xlsx-file

        Arguments:
            message -- an email.message.Message - object

        Returns: tuple
            employee_name -- searched name
            filename -- filename of the attachment
        """

        employee_name = None
        filename = None

        if message.is_multipart():
            for part in message.walk():
                content_type: str = part.get_content_type()
                content_disposition: str = str(part.get("Content-Disposition"))
                try:
                    raw_body : str = part.get_payload(decode=True).decode()
                    body: list = raw_body.splitlines()
                    if employee_name is None:
                        for line in body:
                            if content_type == "text/html" and \
                                "attachment" not in content_disposition:
                                employee_name = parse_html_body(line)
                                if employee_name: break

                    elif content_type == "text/plain" and "attachment" not in content_disposition:
                        if line.strip():
                            employee_name : str = line.strip()
                            break

                except:
                    raw_body = None

                if "attachment" in content_disposition and filename is None:
                    filename: str = part.get_filename()
                    if filename.endswith(".xlsx"):
                        get_attachment(part, filename)
                    else:
                        filename = None

        return employee_name, filename

    def disassemble_sender(sender:str)-> tuple:
        """disassembles the sender-string to name and adress

        Arguments:
            sender -- format: name <name@adress>

        Returns:tuple
            name -- linked name to the email-adress
            adress -- the email-adress
        """

        string :str = sender.split(" ")
        for value in string:
            if "@" in value:
                adress : str = value.strip("<>")
                string.remove(value)
        name : str = ""
        for value in string:
            name += f" {value}"
        name = name.strip(" ")

        return name, adress

    with open(f"{MAILBOX_PATH}{mail}", "r") as file:
        message = email.message_from_file(file)

    sender: str = header.decode_header(message.get("From"))[0][0]
    if message["Subject"] is not None:
        subject: str = header.decode_header(message["Subject"])[0][0]
    else:
        subject = "None"



    if check_subject(subject):
        sender_name, adress = disassemble_sender(sender)
        employee_name, filename = get_content(message)
        if employee_name is None and sender_name is not None:
            employee_name = sender_name
        else:
            adress = sender

    else:
        sender_name, adress = disassemble_sender(sender)
        employee_name, filename = None, None

    return adress, employee_name, filename

def get_attachment(part: email, filename: str)-> None:
    """extracts,decodes and save the attachment from mailpart and rename/move the old one

    Arguments:
        part -- attachmentpart from mail
        filename -- name attachmentfile
    """

    roster_folder: str = f"{OWN_PATH}Schichtpläne/"
    old_roster_folder: str = f"{OWN_PATH}Schichtpläne/Schichtpläne_alt/"

    # move the old file if exsist
    for file in os.listdir(roster_folder):
        if os.path.isfile(f"{roster_folder}{file}"):
            os.rename(f"{roster_folder}{file}", f"{old_roster_folder}{file}")

    open(f"{roster_folder}{filename}", "wb").write(part.get_payload(decode=True))

def extract_data(employee_name:str, filename:str)-> list:
    """extracts data of interest from *.xlsx-file

    At first it searchs column with the names, compares with the given name to finds the
    row with dates. Then it links the dates to the summary and creates an eventlist.

    Arguments:
        employee_name -- searched name
        filename -- filename of the attachment

    Returns: list
        begin -- begin of the shift
        end -- end of the shift
        summary -- summary of the shift
        busy -- True then busy for this event
        all_day -- True if it's an all-day event
    """

    def search_daterow(max_row: int, max_col: int)-> tuple:
        """searchs for the column and row where the date-row starts

        Arguments:
            max_row -- last used row
            max_col -- last used column

        Returns: tuple
            row -- startrow with dates
            col -- startcolumn with dates
        """

        for col in range(1, max_col + 1):
            for row in range(1, max_row + 1):
                cell_obj = sheet.cell(row = row, column = col)
                if isinstance(cell_obj.value, dt.datetime):
                    return row, col

        return None, None

    def search_name(max_row: int, max_col: int, employee_name: str)->tuple:
        """search for the column and row where the name-row is
        comparison the employee-name with the names in the name-column
        if no match found it repeat and searchs for a similar name ("Levenshtein Distance")

        Arguments:
            max_row -- last used row
            max_col -- last used column
            employee_name -- searched name

        Returns: tuple
            row -- row of the first shift-time
            col -- column of first the shift-time
        """

        def check_for_name(searched_name: str, name_list: list)-> bool:
            """search for the name wich is most similar to the searched name

            Arguments:
                searched_name -- the employee-name
                name_list -- the saved list from first turn

            Returns:
                name -- the most similar name
                False -- if there no similar name
            """
            distancelist: list = [Levenshtein.distance(searched_name, name) for name in name_list]
            name: list = name_list[distancelist.index(min(distancelist))]
            if min(distancelist)<5:
                return name
            else:
                return False

        name_list : list = []
        # correct written name
        for col in range(1, max_col):
            for row in range(1, max_row):
                cell_value = sheet.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    name = re.sub(r"\([^()]*\)", " ", cell_value).strip(" ")
                    if name != "": name_list.append(name)
                    if employee_name.lower() == name.lower():
                        return row, col
        # not correct written name
        else:
            name = check_for_name(employee_name, name_list)
            if name:
                return search_name(max_row, max_col, name)  # try again with the most similar name
            else:
                return None, None

    def get_dates(date_row: int, date_col: int, name_row: int)-> list:
        """extract the dates and the belonging times

        Arguments:
            date_row -- row of dates
            date_col -- column where the dates start
            name_row -- row of the times

        Returns: list
            date_time_list -- list of days [date:datetime, period:str or task:str ]
        """

        date_time_list:list = []
        for col in range(date_col, max_col):
            cell_value = sheet.cell(row = date_row, column = col).value
            if isinstance(cell_value, dt.datetime):
                day:list = []
                day.append(cell_value)
                task: str = sheet.cell(row = name_row, column = col).value
                day.append(task)
                if task is not None:
                    date_time_list.append(day)
        return date_time_list

    def convert_datelist(date_time_list:list)-> list:
        """prepares the data for the ics-creator
        - create the correct datetime-objects,
        - link the task to the summary
        - depending on the task set the day as busy respectively as all-day

        Arguments:
            date_time_list -- list of days [date:datetime, period:str or task:str ]

        Returns: list
            duty_roster -- list of the individual day-tuple (summary:str ,begin ,end, busy, all_day)
        """

        # templates for the different kind of shifts
        template_dict : dict = {"seminar": (dt.time(8,0),dt.time(21,0),True,False),
                        "gz früh": (dt.time(7,0),dt.time(14,0),True,False),
                        "gz mittel": (dt.time(12,0),dt.time(19,0),True,False),
                        "gz spät": (dt.time(14,0),dt.time(21,0),True,False),
                        "workday": (dt.time(0,0),dt.time(0,0),True,False),
                        "urlaub":(dt.time(0,0),dt.time(0,0),False,True),
                        "frei":(dt.time(0,0),dt.time(0,0),False,True),
                        "wunschfrei":(dt.time(0,0),dt.time(0,0),False,True),
                        "elternzeit":(dt.time(0,0),dt.time(0,0),False,True)}
                        #"":(dt.time(0,0),dt.time(0,0),False,True)}             # ToDo maybe

        # list of the individual days
        duty_roster = []

        for i in date_time_list:
            date, summary = i
            key = summary.lower().strip(" ").split("-")
            # assign the the summary to a template
            if len(key) > 1:
                summary = "Arbeit"
                begin, end = key
                begin  = dt.datetime.strptime(begin, "%H:%M").time()
                end = dt.datetime.strptime(end, "%H:%M").time()
                busy, all_day = True, False
            elif summary in template_dict:
                begin, end, busy, all_day = template_dict[summary]
            elif summary == "":
                continue
            else:
                begin, end, busy, all_day = template_dict["seminar"]


            if all_day:
                begin = date
                end = begin + dt.timedelta(days=1)
            else:
                begin = dt.datetime.combine(date,begin)
                end = dt.datetime.combine(date,end)

            day = (summary,begin,end, busy, all_day)

            duty_roster.append(day)

        return duty_roster

    if filename is None:
        for f in os.listdir(f"{OWN_PATH}Schichtpläne"):
            if f.endswith(".xlsx"):
                filename = f
                break

    roster = openpyxl.load_workbook(f"{OWN_PATH}Schichtpläne/{filename}", read_only=True)
    sheet = roster.active

    max_row, max_col = sheet.max_row, sheet.max_column
    if max_row > 100 and max_col > 100: max_row, max_col = 100, 100  # ToDo some of my xlsx files have over 400 empty rows or columns

    date_row, date_col = search_daterow(max_row, max_col)
    name_row, name_col = search_name(max_row, max_col, employee_name)
    if name_row is None: return None

    date_time_list = get_dates(date_row, date_col, name_row)
    duty_roster = convert_datelist(date_time_list)

    return duty_roster

def create_ics(employee_name: str, duty_roster: list)-> None:
    """creates an *.ics-file from the extracted data with the employeename

        Arguments:
            employee_name -- name of the searched employee

            duty_roster -- list of the individual day-tuple (summary:str ,begin ,end, busy, all_day)
    """

    duty_roster_cal = ics.Calendar("Schichtplan")

    for entry in duty_roster:
        summary, begin, end, busy, all_day = entry
        shift = ics.Event(organizer="Schichtplan",summary=summary,description=summary,begin=begin,end=end,busy=busy,all_day=all_day,country_code="DE")
        if busy:
            shift.add_alarm(trigger=dt.timedelta(hours=2), description="erste Erinnerung")
        duty_roster_cal.add_event(shift)

    duty_roster_cal.save_ics(filename=f"{OWN_PATH}temp/{employee_name}.ics")

def send_answer(sender:str , employee_name: str, error: int)-> None:
    """send_answer sends the matching answer via E-Mail. If it has worked the *.ics,
    otherwise an E-Mail with an hint.

        Arguments:
            sender -- emailadress of the sender
            employee_name -- searched name
            error -- kind of error
    """

    if error == 0:
        message = """im Anhang befindet sich die gewünschte .ics -Datei mit deinem Namen.
Einfach öffnen und in eine Kalender-App importieren.

Bitte prüfe nochmal ob die Zeiten in dem erstellten Kalender wirklich stimmen. Vor allem für GZ, ÜF, Seminar, etc.
Da sind im Dienstplan leider keine Zeiten hinterlegt"""

    elif error == 1:
        message = """tut mir leid, da is leider etwas schief gegangen.
Vielleicht konnte der Name nicht gefunden, oder der Btreff nicht zugeordnet werden.

Der Name muss in der ersten Zeile der E-mail stehen, am besten genauso wie er im Dienstplan steht und
im Betreff muss "Dienstplan" oder "Schichtplan" stehen
Falls ein anderer Fehler vorliegt werde ich mich schnellstmöglich darum kümmern. ;)"""


    body = f"""Hi,
{message}


Gruß Norman




Bei Problemen oder Fragen bitte eine Mail an:

norman@nortan-ds.de


Die verarbeiteten Daten werden weder gespeichert noch anderweitig weiter genutzt."""

    mail = MIMEMultipart()
    mail["From"] = SERVICEADRESS
    mail["To"] = sender
    mail["Subject"] = "Dienstplan"

    mail.attach(MIMEText(body, "plain"))

    if employee_name and error == 0:
        with open(f"{OWN_PATH}temp/{employee_name}.ics", "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header("Content-Disposition",f"attachment; filename= {employee_name}.ics",)
        mail.attach(part)

    text = mail.as_string()

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP_SSL(SMTP, 465, context=context) as server:
            server.login(EMAILADRESS, PASSWORD)
            server.sendmail(EMAILADRESS, sender, text)
            time.sleep(2)
    except Exception:
        time.sleep(60)

def cleanup(employee_name:str, mail: str, error: int)-> None:
    """deletes the handled mail and the *.ics-file and seperates mails which cause an error

    Arguments:
        employee_name -- searched name
        mail -- filename of the mail
        error -- kind of error
    """
    if error:
        os.rename(f"{MAILBOX_PATH}{mail}", f"{FAULTY_MAILS}{mail}")
    else:
        os.remove(f"{OWN_PATH}temp/{employee_name}.ics")
        os.remove(f"{MAILBOX_PATH}{mail}")

if __name__ == "__main__":

    while True:
        error=0
        new_mails: list = check_mailbox()
        if new_mails:
            for mail in new_mails:
                sender, employee_name, filename = read_mail(mail)
                if employee_name:
                    duty_roster: list = extract_data(employee_name, filename)
                    if duty_roster:
                        create_ics(employee_name, duty_roster)
                    else:
                        error = 1
                else:
                    error = 1

                send_answer(sender, employee_name, error)
                cleanup(employee_name, mail, error)

        time.sleep(60)
